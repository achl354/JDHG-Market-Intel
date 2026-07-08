from __future__ import annotations

import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(__file__).resolve().parents[1]
CONFIG = ROOT / "config"
REPORTS = ROOT / "reports"
DATA = ROOT / "data"
SNAPSHOT_FILE = DATA / "source_snapshots" / "source_hashes.json"
FINDINGS_XLSX = DATA / "findings.xlsx"
SYDNEY = ZoneInfo("Australia/Sydney")

PRODUCT_TERMS = [
    "patient handling",
    "patient positioning",
    "positioning",
    "lateral transfer",
    "air-assisted",
    "air assisted",
    "floor recovery",
    "hoist",
    "lifter",
    "sling",
    "transfer board",
    "slide sheet",
    "pressure care",
    "pressure injury",
    "hospital bed",
    "bed mover",
    "stretcher",
    "trolley",
    "treatment chair",
    "transfer chair",
    "underpad",
    "apron",
    "gown",
    "ppe",
    "curtain",
    "infection control",
    "theatre",
    "surgical",
]

CHANGE_TERMS = [
    "new",
    "launch",
    "launched",
    "introducing",
    "available",
    "catalogue",
    "brochure",
    "product alert",
    "recall",
    "safety",
    "contract",
    "partnership",
    "award",
    "artg",
    "registered",
]

ACCOUNT_TERMS = [
    "calvary",
    "healthscope",
    "ramsay",
    "st john of god",
    "mater",
    "epworth",
    "cabrini",
    "princess alexandra",
    "pa hospital",
    "monash health",
    "austin health",
]


@dataclass
class Candidate:
    source: str
    source_type: str
    url: str
    date_found: str
    competitor_or_sponsor: str
    product_category: str
    title: str
    snippet: str
    evidence: str


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def should_run(now: datetime) -> bool:
    if os.getenv("FORCE_RUN", "").lower() == "true":
        return True
    local = now.astimezone(SYDNEY)
    anchor = datetime(2026, 7, 20, 8, 0, tzinfo=SYDNEY)
    if local.weekday() != 0 or local.hour != 8:
        print(f"Skipping: local Sydney time is {local.isoformat()}, not Monday 8am.")
        return False
    days = (local.date() - anchor.date()).days
    if days < 0 or days % 14 != 0:
        print(f"Skipping: {local.date()} is not on the fortnightly cadence from {anchor.date()}.")
        return False
    return True


def slugify(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "scan"


def clean_text(value: str, limit: int = 900) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    return value[:limit]


def http_get(url: str) -> str:
    headers = {"User-Agent": "JDHG-Market-Intel/1.0 (+https://github.com/achl354/JDHG-Market-Intel)"}
    response = requests.get(url, headers=headers, timeout=25)
    response.raise_for_status()
    return response.text


def page_summary(url: str) -> tuple[str, str, str]:
    try:
        html = http_get(url)
    except Exception as exc:
        return "Fetch failed", f"Could not fetch source page: {exc}", ""
    soup = BeautifulSoup(html, "html.parser")
    title = clean_text(soup.title.get_text(" ")) if soup.title else url
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    body = clean_text(soup.get_text(" "), 1600)
    digest = hashlib.sha256(body.encode("utf-8", errors="ignore")).hexdigest()
    return title, body, digest


def tavily_search(query: str, max_results: int = 5) -> list[dict[str, Any]]:
    api_key = os.getenv("TAVILY_API_KEY")
    if not api_key:
        raise RuntimeError("TAVILY_API_KEY is not set.")
    payload = {
        "api_key": api_key,
        "query": query,
        "search_depth": "basic",
        "max_results": max_results,
        "include_answer": False,
        "include_raw_content": False,
    }
    response = requests.post("https://api.tavily.com/search", json=payload, timeout=40)
    response.raise_for_status()
    return response.json().get("results", [])


def collect_candidates(watchlist: dict[str, Any], artg_config: dict[str, Any], date_found: str) -> list[Candidate]:
    candidates: list[Candidate] = []

    for company in watchlist["competitors"]:
        for query in company.get("queries", [])[:3]:
            for result in tavily_search(query, max_results=4):
                candidates.append(
                    Candidate(
                        source=result.get("url", ""),
                        source_type="Competitor website / web search",
                        url=result.get("url", ""),
                        date_found=date_found,
                        competitor_or_sponsor=company["name"],
                        product_category="; ".join(watchlist["jdhg_categories"][:4]),
                        title=clean_text(result.get("title", "")),
                        snippet=clean_text(result.get("content", "")),
                        evidence=f"Search query: {query}",
                    )
                )

    for supplier in watchlist["jdhg_suppliers_to_contextualise"]:
        for query in supplier.get("competitor_queries", [])[:2]:
            for result in tavily_search(query, max_results=3):
                candidates.append(
                    Candidate(
                        source=result.get("url", ""),
                        source_type="Supplier competitor ecosystem",
                        url=result.get("url", ""),
                        date_found=date_found,
                        competitor_or_sponsor=supplier["name"],
                        product_category="Supplier-adjacent competitor activity",
                        title=clean_text(result.get("title", "")),
                        snippet=clean_text(result.get("content", "")),
                        evidence=f"Search query: {query}",
                    )
                )

    for source in artg_config["artg_sources"]:
        for query in source.get("queries", []):
            for result in tavily_search(query, max_results=4):
                candidates.append(
                    Candidate(
                        source=result.get("url", ""),
                        source_type="ARTG / regulatory",
                        url=result.get("url", ""),
                        date_found=date_found,
                        competitor_or_sponsor="TGA / ARTG",
                        product_category="Regulatory signal",
                        title=clean_text(result.get("title", "")),
                        snippet=clean_text(result.get("content", "")),
                        evidence=f"Search query: {query}",
                    )
                )

    candidates.extend(detect_source_page_changes(watchlist, date_found))
    return dedupe_candidates(candidates)


def detect_source_page_changes(watchlist: dict[str, Any], date_found: str) -> list[Candidate]:
    previous = {}
    if SNAPSHOT_FILE.exists():
        previous = load_json(SNAPSHOT_FILE)

    next_snapshot: dict[str, str] = {}
    candidates: list[Candidate] = []
    entities = watchlist["competitors"] + watchlist["jdhg_suppliers_to_contextualise"]

    for entity in entities:
        url = entity.get("website")
        if not url:
            continue
        title, body, digest = page_summary(url)
        if digest:
            next_snapshot[url] = digest
        old_digest = previous.get(url)
        if old_digest and digest and old_digest != digest:
            candidates.append(
                Candidate(
                    source=url,
                    source_type="Website change detection",
                    url=url,
                    date_found=date_found,
                    competitor_or_sponsor=entity["name"],
                    product_category="Website / catalogue / resource update",
                    title=title,
                    snippet=body,
                    evidence="Homepage/body text hash changed since previous scan.",
                )
            )

    SNAPSHOT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with SNAPSHOT_FILE.open("w", encoding="utf-8") as f:
        json.dump(next_snapshot, f, indent=2, sort_keys=True)
    return candidates


def dedupe_candidates(candidates: list[Candidate]) -> list[Candidate]:
    seen = set()
    deduped = []
    for item in candidates:
        key = (item.url, item.title, item.competitor_or_sponsor)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped[:80]


def classify_with_claude(candidates: list[Candidate], profile: dict[str, Any], watchlist: dict[str, Any]) -> list[dict[str, Any]]:
    if not candidates:
        return []
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return heuristic_findings(candidates, "Claude unavailable: ANTHROPIC_API_KEY not set")

    compact = [
        {
            "source": c.source,
            "source_type": c.source_type,
            "url": c.url,
            "date_found": c.date_found,
            "competitor_or_sponsor": c.competitor_or_sponsor,
            "product_category": c.product_category,
            "title": c.title,
            "snippet": c.snippet,
            "evidence": c.evidence,
        }
        for c in candidates[:50]
    ]
    prompt = {
        "role": "user",
        "content": (
            "You are producing competitor intelligence for JD Healthcare Group in Australia. "
            "Classify only genuinely relevant findings. Do not overstate claims. If evidence is weak, say it is a possible signal. "
            "Return strict JSON only, with a top-level key 'findings'. Each finding must include: "
            "source, date_found, competitor_or_sponsor, product_category, what_changed, why_it_matters, "
            "priority, suggested_jdhg_action, confidence, source_type, url. "
            f"Allowed actions: {profile['action_labels']}. "
            "Priority must be High priority, Medium priority, or Low priority. "
            "Confidence must be High, Medium, or Low, with a short note after a colon if useful. "
            f"JDHG categories: {watchlist['jdhg_categories']}. "
            f"Candidates: {json.dumps(compact, ensure_ascii=False)}"
        ),
    }
    base_payload = {
        "max_tokens": 6000,
        "temperature": 0.1,
        "messages": [prompt],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    preferred_model = os.getenv("ANTHROPIC_MODEL", "claude-3-5-haiku-latest")
    fallback_models = [
        preferred_model,
        "claude-3-5-haiku-20241022",
        "claude-3-haiku-20240307",
    ]
    tried: list[str] = []
    for model in dict.fromkeys(fallback_models):
        tried.append(model)
        payload = {"model": model, **base_payload}
        response = requests.post("https://api.anthropic.com/v1/messages", headers=headers, json=payload, timeout=90)
        if response.ok:
            data = response.json()
            text = "".join(block.get("text", "") for block in data.get("content", []) if block.get("type") == "text")
            return parse_findings_json(text, candidates)
        print(f"Claude request failed for model {model}: HTTP {response.status_code} {response.text[:1000]}")

    note = f"Claude API failed for models: {', '.join(tried)}"
    return heuristic_findings(candidates, note)


def parse_findings_json(text: str, candidates: list[Candidate]) -> list[dict[str, Any]]:
    match = re.search(r"\{.*\}", text, flags=re.S)
    if not match:
        return heuristic_findings(candidates, "Claude response was not parseable JSON")
    parsed = json.loads(match.group(0))
    findings = parsed.get("findings", [])
    cleaned = []
    for finding in findings:
        if not finding.get("what_changed") or finding.get("suggested_jdhg_action") == "No action":
            continue
        cleaned.append(finding)
    return cleaned[:35]


def heuristic_findings(candidates: list[Candidate], note: str, limit: int = 30) -> list[dict[str, Any]]:
    ranked = sorted(candidates, key=heuristic_score, reverse=True)
    findings = []
    seen_urls = set()
    for candidate in ranked:
        if candidate.url in seen_urls:
            continue
        seen_urls.add(candidate.url)
        score = heuristic_score(candidate)
        if score < 3:
            continue
        findings.append(heuristic_finding(candidate, note))
        if len(findings) >= limit:
            break
    if findings:
        return findings
    return [heuristic_finding(c, note) for c in ranked[: min(limit, len(ranked))]]


def heuristic_score(candidate: Candidate) -> int:
    text = f"{candidate.title} {candidate.snippet} {candidate.evidence} {candidate.source_type}".lower()
    score = 0
    score += sum(1 for term in PRODUCT_TERMS if term in text)
    score += 2 * sum(1 for term in CHANGE_TERMS if term in text)
    score += 2 * sum(1 for term in ACCOUNT_TERMS if term in text)
    if "haines" in candidate.competitor_or_sponsor.lower():
        score += 4
    if "artg" in text or "regulatory" in candidate.source_type.lower() or "tga.gov.au" in candidate.url:
        score += 5
    if "linkedin.com" in candidate.url:
        score += 1
    if "news" in candidate.url or "blog" in candidate.url or "resource" in candidate.url:
        score += 2
    return score


def heuristic_finding(candidate: Candidate, note: str) -> dict[str, Any]:
    text = f"{candidate.title} {candidate.snippet} {candidate.evidence} {candidate.url}".lower()
    score = heuristic_score(candidate)
    priority = "Low priority"
    action = choose_action(candidate, text)
    if score >= 12:
        priority = "High priority"
    elif score >= 6:
        priority = "Medium priority"
    if "haines" in candidate.competitor_or_sponsor.lower():
        priority = "High priority" if score >= 8 else "Medium priority"
    if any(term in text for term in ["recall", "safety alert", "product alert"]):
        priority = "High priority"

    matched_terms = [term for term in PRODUCT_TERMS + CHANGE_TERMS + ACCOUNT_TERMS if term in text][:8]
    what_changed = describe_change(candidate, matched_terms)
    why_it_matters = describe_relevance(candidate, matched_terms, text)
    return {
        "source": candidate.source,
        "date_found": candidate.date_found,
        "competitor_or_sponsor": candidate.competitor_or_sponsor,
        "product_category": candidate.product_category,
        "what_changed": what_changed,
        "why_it_matters": why_it_matters,
        "priority": priority,
        "suggested_jdhg_action": action,
        "confidence": f"{heuristic_confidence(score)}: Rule-based triage because Claude did not run. {note}",
        "source_type": candidate.source_type,
        "url": candidate.url,
    }


def choose_action(candidate: Candidate, text: str) -> str:
    if candidate.source_type.startswith("ARTG") or "artg" in text or "tga.gov.au" in candidate.url:
        return "Compliance review"
    if any(term in text for term in ["calvary", "healthscope", "ramsay", "cabrini", "mater", "epworth"]):
        return "Account follow-up"
    if any(term in text for term in ["catalogue", "brochure", "resource", "white paper", "education"]):
        return "Update sales collateral"
    if any(term in text for term in ["new", "launch", "introducing", "available", "product alert"]):
        return "Review product overlap"
    if "contract" in text or "tender" in text or "panel" in text:
        return "Add to tender intelligence"
    return "Monitor"


def describe_change(candidate: Candidate, matched_terms: list[str]) -> str:
    title = candidate.title or "Untitled source"
    if candidate.source_type.startswith("ARTG"):
        return f"Possible ARTG/regulatory signal surfaced for review: {title}"
    if candidate.source_type == "Website change detection":
        return f"Website content changed since the previous scan: {title}"
    if "Supplier competitor" in candidate.source_type:
        return f"Supplier-ecosystem competitor signal surfaced: {title}"
    if matched_terms:
        return f"Possible relevant market signal surfaced ({', '.join(matched_terms[:5])}): {title}"
    return f"Possible market signal surfaced: {title}"


def describe_relevance(candidate: Candidate, matched_terms: list[str], text: str) -> str:
    reasons = []
    if "haines" in candidate.competitor_or_sponsor.lower():
        reasons.append("Haines is a priority competitor for JDHG.")
    if candidate.source_type.startswith("ARTG") or "artg" in text:
        reasons.append("Regulatory entries can indicate new supply capability, sponsor activity, or category compliance changes.")
    if any(term in text for term in ["new", "launch", "introducing", "available", "product alert"]):
        reasons.append("The wording suggests a product, catalogue, or availability change that should be verified.")
    if any(term in text for term in ACCOUNT_TERMS):
        reasons.append("The signal mentions a watched hospital or private hospital group.")
    if matched_terms:
        reasons.append(f"Matched JDHG-relevant terms: {', '.join(matched_terms[:8])}.")
    snippet = clean_text(candidate.snippet, 360)
    if snippet:
        reasons.append(f"Evidence snippet: {snippet}")
    return " ".join(reasons) or "The source was returned by a targeted JDHG market-watch query and should be reviewed for relevance."


def heuristic_confidence(score: int) -> str:
    if score >= 12:
        return "Medium"
    if score >= 6:
        return "Low-Medium"
    return "Low"


def render_markdown(findings: list[dict[str, Any]], profile: dict[str, Any], date_found: str) -> str:
    grouped = {
        "High priority": [f for f in findings if f.get("priority") == "High priority"],
        "Medium priority": [f for f in findings if f.get("priority") == "Medium priority"],
        "Low priority": [f for f in findings if f.get("priority") == "Low priority"],
    }
    lines = [
        f"# {profile['report_title']}",
        "",
        f"**Date found:** {date_found}",
        f"**Region focus:** {profile['region_focus']}",
        "",
        "## 1. Executive Summary",
        "",
        f"- {len(findings)} relevant findings were identified for review.",
        f"- High priority: {len(grouped['High priority'])}",
        f"- Medium priority: {len(grouped['Medium priority'])}",
        f"- Low priority: {len(grouped['Low priority'])}",
        "- Tender monitoring is currently excluded from this workflow.",
        "",
    ]
    section_map = [
        ("## 2. High-Priority Signals", grouped["High priority"]),
        ("## 3. Medium-Priority Signals", grouped["Medium priority"]),
        ("## 4. Low-Priority Signals", grouped["Low priority"]),
        ("## 5. New ARTG / Regulatory Signals", [f for f in findings if "ARTG" in f.get("source_type", "") or "regulatory" in f.get("source_type", "").lower()]),
        ("## 6. Competitor Website / Announcement Changes", [f for f in findings if "Website" in f.get("source_type", "") or "Competitor" in f.get("source_type", "")]),
        ("## 7. LinkedIn / Market Messaging Signals", [f for f in findings if "linkedin" in f.get("source", "").lower() or "linkedin" in f.get("url", "").lower()]),
        ("## 8. Tender Award / Contract Signals", []),
        ("## 9. Suggested JDHG Actions", findings),
        ("## 10. Watchlist for Next Scan", []),
    ]
    for heading, items in section_map:
        lines.extend([heading, ""])
        if heading.endswith("Tender Award / Contract Signals"):
            lines.extend(["Tenders are intentionally excluded from this version.", ""])
            continue
        if heading.endswith("Watchlist for Next Scan"):
            lines.extend(["- Confirm any low-confidence LinkedIn/company-page matches.", "- Watch for follow-on product, catalogue or ARTG updates from High/Medium findings.", ""])
            continue
        if not items:
            lines.extend(["No relevant signals found in this scan.", ""])
            continue
        for f in items:
            lines.extend(format_finding_md(f))
    return "\n".join(lines)


def format_finding_md(finding: dict[str, Any]) -> list[str]:
    return [
        f"### {finding.get('competitor_or_sponsor', 'Unknown')} - {finding.get('product_category', 'General')}",
        "",
        f"- **Source:** [{finding.get('source', finding.get('url', 'Source'))}]({finding.get('url') or finding.get('source')})",
        f"- **Date found:** {finding.get('date_found', '')}",
        f"- **Priority:** {finding.get('priority', '')}",
        f"- **What changed:** {finding.get('what_changed', '')}",
        f"- **Why it matters:** {finding.get('why_it_matters', '')}",
        f"- **Suggested JDHG action:** {finding.get('suggested_jdhg_action', '')}",
        f"- **Confidence:** {finding.get('confidence', '')}",
        "",
    ]


def update_excel(findings: list[dict[str, Any]]) -> None:
    headers = [
        "date_found",
        "priority",
        "competitor_or_sponsor",
        "product_category",
        "what_changed",
        "why_it_matters",
        "suggested_jdhg_action",
        "confidence",
        "source_type",
        "source",
        "url",
    ]
    if FINDINGS_XLSX.exists():
        wb = load_workbook(FINDINGS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Findings"
        ws.append(headers)
    existing = {row[10].value for row in ws.iter_rows(min_row=2) if row[10].value}
    for finding in findings:
        if finding.get("url") in existing:
            continue
        ws.append([finding.get(h, "") for h in headers])
    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 55)
    DATA.mkdir(exist_ok=True)
    wb.save(FINDINGS_XLSX)


def render_pdf(findings: list[dict[str, Any]], pdf_path: Path, title: str, date_found: str) -> None:
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="JDHGTitle", parent=styles["Title"], textColor=colors.HexColor("#143D59"), fontSize=20, leading=24))
    styles.add(ParagraphStyle(name="JDHGHeading", parent=styles["Heading2"], textColor=colors.HexColor("#0B5C6B"), fontSize=13, leading=16))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10))
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4, rightMargin=16 * mm, leftMargin=16 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    story: list[Any] = [
        Paragraph(title, styles["JDHGTitle"]),
        Paragraph(f"Date found: {date_found}", styles["BodyText"]),
        Spacer(1, 6 * mm),
        Paragraph("Executive Summary", styles["JDHGHeading"]),
        Paragraph(f"{len(findings)} relevant findings identified. Tenders are excluded from this version.", styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]
    for priority in ["High priority", "Medium priority", "Low priority"]:
        items = [f for f in findings if f.get("priority") == priority]
        story.append(Paragraph(priority, styles["JDHGHeading"]))
        if not items:
            story.append(Paragraph("No relevant signals found.", styles["BodyText"]))
            continue
        for f in items[:12]:
            data = [
                ["Competitor / sponsor", clean_text(f.get("competitor_or_sponsor", ""), 120)],
                ["Product / category", clean_text(f.get("product_category", ""), 120)],
                ["What changed", clean_text(f.get("what_changed", ""), 300)],
                ["Why it matters", clean_text(f.get("why_it_matters", ""), 300)],
                ["Action", clean_text(f.get("suggested_jdhg_action", ""), 120)],
                ["Confidence", clean_text(f.get("confidence", ""), 120)],
                ["Source", clean_text(f.get("url") or f.get("source", ""), 150)],
            ]
            table = Table(data, colWidths=[38 * mm, 124 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F1F2")),
                        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#143D59")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C7CC")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.extend([table, Spacer(1, 4 * mm)])
    doc.build(story)


def main() -> int:
    now = datetime.now(timezone.utc)
    if not should_run(now):
        return 0
    REPORTS.mkdir(exist_ok=True)
    DATA.mkdir(exist_ok=True)

    profile = load_json(CONFIG / "report_profile.json")
    watchlist = load_json(CONFIG / "watchlist.json")
    artg_config = load_json(CONFIG / "artg_keywords.json")
    date_found = now.astimezone(SYDNEY).strftime("%Y-%m-%d")

    candidates = collect_candidates(watchlist, artg_config, date_found)
    findings = classify_with_claude(candidates, profile, watchlist)

    md = render_markdown(findings, profile, date_found)
    report_base = f"{date_found}-jdhg-market-landscape"
    md_path = REPORTS / f"{report_base}.md"
    pdf_path = REPORTS / f"{report_base}.pdf"
    md_path.write_text(md, encoding="utf-8")
    render_pdf(findings, pdf_path, profile["report_title"], date_found)
    update_excel(findings)

    print(f"Wrote {md_path}")
    print(f"Wrote {pdf_path}")
    print(f"Updated {FINDINGS_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
